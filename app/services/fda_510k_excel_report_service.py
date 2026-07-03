"""FDA 510(k) final SBOM Excel report generation from persisted data."""

from __future__ import annotations

import json
import logging
import re
from copy import copy
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..metrics import COMPLETED_RUN_STATUSES
from ..models import AnalysisFinding, AnalysisRun, Projects, SBOMComponent, SBOMSource, VexStatement

log = logging.getLogger(__name__)

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FDA_510K_TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "templates" / "reports" / "FDA_510k_SBOM_Template 1.xlsx"

DATA_SHEETS = ("SBOM Components", "Vulnerabilities & VEX", "Lifecycle & Support Plan")
EXPECTED_SHEETS = ("Instructions", "SBOM Metadata", *DATA_SHEETS)
INCOMPLETE_ANALYSIS_CODE = "fda_510k_report_incomplete_analysis"


class Fda510kReportError(ValueError):
    """Base class for request validation errors."""


class Fda510kTemplateMissingError(Fda510kReportError):
    """Raised when the approved workbook template is unavailable."""


class Fda510kIncompleteAnalysisError(Fda510kReportError):
    """Raised when required persisted analyses are not complete."""

    def __init__(self, blockers: list[dict[str, Any]]) -> None:
        self.blockers = blockers
        super().__init__("Required analyses are incomplete for one or more selected SBOMs.")

    def detail(self) -> dict[str, Any]:
        return {
            "code": INCOMPLETE_ANALYSIS_CODE,
            "message": str(self),
            "blockers": self.blockers,
        }


@dataclass(frozen=True)
class Fda510kReportMetadata:
    device_name: str
    manufacturer_sponsor: str
    device_software_version: str
    author_of_sbom_data: str
    prepared_by: str
    device_model_catalog_number: str | None = None
    submission_type: str | None = None
    submission_number: str | None = None
    product_code_regulation_number: str | None = None
    top_level_primary_component: str | None = None
    sbom_version: str | None = None
    sbom_formats_for_submission: str | None = None
    sbom_generation_tool_and_version: str | None = None
    primary_data_source: str | None = None
    date_prepared: date | None = None
    reviewed_approved_by: str | None = None
    date_approved: date | None = None


@dataclass(frozen=True)
class Fda510kSelection:
    sbom_id: int
    findings_analysis_run_id: int | None = None
    lifecycle_analysis_run_id: int | None = None


@dataclass
class _ComponentAggregate:
    key: str
    name: str = ""
    version: str = ""
    supplier: str = ""
    unique_id_type: str = ""
    unique_id_value: str = ""
    component_type: str = ""
    origin_category: str = ""
    license: str = ""
    support_level: str = ""
    eos_date: date | None = None
    eol_date: date | None = None
    patch_mechanism: str = ""
    crypto_info: str = ""
    source_sboms: set[str] = field(default_factory=set)
    dependencies: set[str] = field(default_factory=set)
    vulnerabilities: set[str] = field(default_factory=set)
    lifecycle_recommendations: set[str] = field(default_factory=set)
    components: list[SBOMComponent] = field(default_factory=list)


def _clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def _first_text(*values: Any) -> str:
    return next((_clean(value) for value in values if _clean(value)), "")


def _parse_json(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return value
    if not isinstance(value, str) or not value.strip():
        return None
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return None


def _as_list(value: Any) -> list[str]:
    parsed = _parse_json(value)
    if isinstance(parsed, list):
        return [_clean(item) for item in parsed if _clean(item)]
    if isinstance(value, (list, tuple, set)):
        return [_clean(item) for item in value if _clean(item)]
    text = _clean(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[,;\n]", text) if part.strip()]


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for candidate in (text, text[:10]):
        try:
            return date.fromisoformat(candidate)
        except ValueError:
            continue
    return None


def _severity(value: Any) -> str:
    text = _clean(value).lower()
    mapping = {
        "critical": "Critical",
        "high": "High",
        "medium": "Medium",
        "moderate": "Medium",
        "low": "Low",
        "none": "None",
        "info": "None",
        "informational": "None",
    }
    return mapping.get(text, "None" if not text else text[:1].upper() + text[1:])


def _safe_filename(value: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    safe = safe.strip("._")
    return safe or "project"


def _component_identity(component: SBOMComponent) -> str:
    for prefix, value in (
        ("purl", component.normalized_purl or component.purl),
        ("cpe", component.primary_cpe or component.cpe),
        ("identity", component.normalized_component_key),
        ("canonical", component.dedupe_canonical_id),
    ):
        if _clean(value):
            return f"{prefix}:{_clean(value).casefold()}"
    fallback = "|".join(
        [
            _clean(component.normalized_supplier or component.supplier).casefold(),
            _clean(component.normalized_name or component.name).casefold(),
            _clean(component.normalized_version or component.version).casefold(),
            _clean(component.normalized_ecosystem or component.ecosystem).casefold(),
            _clean(component.purl_namespace or component.component_group).casefold(),
        ]
    )
    return f"fallback:{fallback}"


def _unique_id(component: SBOMComponent) -> tuple[str, str]:
    if _clean(component.normalized_purl or component.purl):
        return "PURL", _clean(component.normalized_purl or component.purl)
    if _clean(component.primary_cpe or component.cpe):
        return "CPE", _clean(component.primary_cpe or component.cpe)
    if _clean(component.bom_ref):
        return "Other", _clean(component.bom_ref)
    return "", ""


def _support_level(component: SBOMComponent) -> str:
    status = _first_text(component.maintenance_status, component.lifecycle_status)
    lowered = status.casefold()
    if lowered in {"active", "supported", "maintained", "eol soon"}:
        return "Actively maintained"
    if "deprecated" in lowered or "unmaintained" in lowered:
        return "No longer maintained"
    if "abandoned" in lowered or "unsupported" in lowered or "eol" in lowered or "eos" in lowered:
        return "Abandoned" if "abandoned" in lowered else "No longer maintained"
    return "Unknown" if not status else status


def _origin_category(component: SBOMComponent) -> str:
    license_text = _clean(component.license).casefold()
    if any(token in license_text for token in ("mit", "apache", "gpl", "bsd", "mpl", "epl")):
        return "Open-Source (OSS)"
    if _clean(component.supplier):
        return "Commercial (COTS)"
    return "Other"


def _risk_for_lifecycle(component: SBOMComponent) -> str:
    status = _clean(component.lifecycle_status).casefold()
    if status in {"eol", "eos", "eof", "unsupported"} or component.unsupported:
        return "High"
    if "deprecated" in status or "soon" in status or component.deprecated or component.is_deprecated:
        return "Medium"
    return "Low"


def _extract_dependency_map(sboms: list[SBOMSource]) -> dict[tuple[int, str], set[str]]:
    result: dict[tuple[int, str], set[str]] = {}
    for sbom in sboms:
        raw = _parse_json(sbom.sbom_data)
        if not isinstance(raw, dict):
            continue
        names: dict[str, str] = {}
        for component in raw.get("components") or raw.get("packages") or []:
            if not isinstance(component, dict):
                continue
            ref = _first_text(component.get("bom-ref"), component.get("SPDXID"), component.get("spdxid"), component.get("id"))
            name = _first_text(component.get("name"), component.get("packageName"), ref)
            if ref:
                names[ref] = name

        for dep in raw.get("dependencies") or []:
            if not isinstance(dep, dict):
                continue
            ref = _clean(dep.get("ref"))
            depends = [_clean(item) for item in dep.get("dependsOn") or [] if _clean(item)]
            if ref and depends:
                result.setdefault((sbom.id, ref), set()).add("Depends on: " + ", ".join(names.get(item, item) for item in depends))
                for item in depends:
                    result.setdefault((sbom.id, item), set()).add(f"Required by: {names.get(ref, ref)}")

        for rel in raw.get("relationships") or []:
            if not isinstance(rel, dict):
                continue
            source = _first_text(rel.get("spdxElementId"), rel.get("source"))
            target = _first_text(rel.get("relatedSpdxElement"), rel.get("target"))
            rel_type = _first_text(rel.get("relationshipType"), rel.get("type"))
            if source and target and rel_type:
                result.setdefault((sbom.id, source), set()).add(f"{rel_type}: {names.get(target, target)}")
    return result


def _copy_row_template(ws: Worksheet, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
        if source.value and isinstance(source.value, str) and source.value.startswith("="):
            target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
        else:
            target.value = None


def _clear_and_prepare_rows(ws: Worksheet, *, start_row: int, style_row: int, last_row: int) -> None:
    for row in range(start_row, max(ws.max_row, last_row) + 1):
        _copy_row_template(ws, style_row, row)


def _set_row_values(ws: Worksheet, row: int, values: dict[int, Any]) -> None:
    for col, value in values.items():
        ws.cell(row, col).value = value


class Fda510kExcelReportService:
    """Build the approved FDA workbook from persisted SBOM analysis data."""

    def __init__(self, db: Session, *, template_path: Path = FDA_510K_TEMPLATE_PATH) -> None:
        self.db = db
        self.template_path = template_path

    def build(
        self,
        project_id: int,
        selections: list[Fda510kSelection],
        metadata: Fda510kReportMetadata,
    ) -> tuple[bytes, str]:
        project, sboms, runs = self._validate_request(project_id, selections)
        dependencies = _extract_dependency_map(sboms)
        component_rows = self._component_aggregates(sboms, runs, dependencies)
        vulnerability_rows = self._vulnerability_rows(runs)
        lifecycle_rows = self._lifecycle_rows(component_rows)
        content = self._build_workbook(metadata, component_rows, vulnerability_rows, lifecycle_rows)
        filename = self.filename_for(project.project_name)
        return content, filename

    @staticmethod
    def filename_for(project_name: str | None) -> str:
        stamp = datetime.now(UTC).strftime("%Y%m%d_%H%M")
        return f"{_safe_filename(project_name or 'project')}_FDA_510k_SBOM_Report_{stamp}.xlsx"

    def _validate_request(
        self,
        project_id: int,
        selections: list[Fda510kSelection],
    ) -> tuple[Projects, list[SBOMSource], dict[int, AnalysisRun]]:
        if not self.template_path.exists():
            raise Fda510kTemplateMissingError(f"Expected FDA 510(k) template at {self.template_path}")
        if not selections:
            raise Fda510kReportError("At least one SBOM must be selected.")

        project = self.db.get(Projects, project_id)
        if project is None or not project.is_active:
            raise Fda510kReportError("Project not found.")

        unique_ids = list(dict.fromkeys(selection.sbom_id for selection in selections))
        if len(unique_ids) != len(selections):
            raise Fda510kReportError("Duplicate SBOM selections are not allowed.")

        sboms = list(
            self.db.execute(select(SBOMSource).where(SBOMSource.id.in_(unique_ids)).order_by(SBOMSource.id.asc())).scalars()
        )
        by_id = {sbom.id: sbom for sbom in sboms}
        missing = [sbom_id for sbom_id in unique_ids if sbom_id not in by_id]
        if missing:
            raise Fda510kReportError(f"Selected SBOM not found: {missing[0]}")

        wrong_project = [sbom for sbom in sboms if sbom.projectid != project_id]
        if wrong_project:
            raise Fda510kReportError(f"Selected SBOM '{wrong_project[0].sbom_name}' does not belong to project {project_id}.")

        blockers: list[dict[str, Any]] = []
        runs: dict[int, AnalysisRun] = {}
        for selection in selections:
            sbom = by_id[selection.sbom_id]
            run = self._resolve_findings_run(sbom, selection.findings_analysis_run_id)
            if run is None:
                blockers.append(self._blocker(sbom, "findings", self._latest_findings_status(sbom)))
            elif run.run_status not in COMPLETED_RUN_STATUSES:
                blockers.append(self._blocker(sbom, "findings", run.run_status or "unknown"))
            else:
                runs[sbom.id] = run

            if selection.lifecycle_analysis_run_id is not None:
                lifecycle_run = self.db.get(AnalysisRun, selection.lifecycle_analysis_run_id)
                if lifecycle_run is None or lifecycle_run.sbom_id != sbom.id:
                    raise Fda510kReportError(
                        f"Lifecycle analysis run {selection.lifecycle_analysis_run_id} does not belong to SBOM {sbom.id}."
                    )
                if lifecycle_run.run_status not in COMPLETED_RUN_STATUSES:
                    blockers.append(self._blocker(sbom, "lifecycle", lifecycle_run.run_status or "unknown"))

            lifecycle_status = self._lifecycle_completion_status(sbom)
            if lifecycle_status != "completed":
                blockers.append(self._blocker(sbom, "lifecycle", lifecycle_status))

        if blockers:
            raise Fda510kIncompleteAnalysisError(blockers)
        return project, [by_id[sbom_id] for sbom_id in unique_ids], runs

    def _resolve_findings_run(self, sbom: SBOMSource, run_id: int | None) -> AnalysisRun | None:
        if run_id is not None:
            run = self.db.get(AnalysisRun, run_id)
            if run is None or run.sbom_id != sbom.id:
                raise Fda510kReportError(f"Findings analysis run {run_id} does not belong to SBOM {sbom.id}.")
            return run
        return self.db.execute(
            select(AnalysisRun)
            .where(AnalysisRun.sbom_id == sbom.id, AnalysisRun.run_status.in_(COMPLETED_RUN_STATUSES))
            .order_by(AnalysisRun.completed_on.desc(), AnalysisRun.id.desc())
            .limit(1)
        ).scalar_one_or_none()

    def _latest_findings_status(self, sbom: SBOMSource) -> str:
        run = self.db.execute(
            select(AnalysisRun)
            .where(AnalysisRun.sbom_id == sbom.id)
            .order_by(AnalysisRun.completed_on.desc(), AnalysisRun.id.desc())
            .limit(1)
        ).scalar_one_or_none()
        return run.run_status if run is not None and run.run_status else "missing"

    def _lifecycle_completion_status(self, sbom: SBOMSource) -> str:
        status = _clean(sbom.enrichment_status).casefold()
        if status in {"running", "pending", "queued", "in_progress"}:
            return "running"
        if status in {"failed", "error", "cancelled", "canceled"}:
            return status
        components = list(
            self.db.execute(
                select(SBOMComponent).where(
                    SBOMComponent.sbom_id == sbom.id,
                    (SBOMComponent.is_duplicate.is_(False)) | (SBOMComponent.is_duplicate.is_(None)),
                )
            ).scalars()
        )
        if not components:
            return "missing"
        if all(_clean(component.lifecycle_checked_at) for component in components):
            return "completed"
        return "missing"

    @staticmethod
    def _blocker(sbom: SBOMSource, analysis_type: str, status: str) -> dict[str, Any]:
        return {
            "sbom_id": sbom.id,
            "sbom_name": sbom.sbom_name,
            "analysis_type": analysis_type,
            "status": status,
        }

    def _component_aggregates(
        self,
        sboms: list[SBOMSource],
        runs: dict[int, AnalysisRun],
        dependencies: dict[tuple[int, str], set[str]],
    ) -> list[_ComponentAggregate]:
        findings_by_component: dict[int, set[str]] = {}
        for finding in self.db.execute(
            select(AnalysisFinding).where(AnalysisFinding.analysis_run_id.in_([run.id for run in runs.values()]))
        ).scalars():
            if finding.component_id is not None:
                findings_by_component.setdefault(finding.component_id, set()).add(finding.vuln_id)

        aggregates: dict[str, _ComponentAggregate] = {}
        components = list(
            self.db.execute(
                select(SBOMComponent)
                .where(
                    SBOMComponent.sbom_id.in_([sbom.id for sbom in sboms]),
                    (SBOMComponent.is_duplicate.is_(False)) | (SBOMComponent.is_duplicate.is_(None)),
                )
                .order_by(SBOMComponent.name.asc(), SBOMComponent.version.asc(), SBOMComponent.id.asc())
            ).scalars()
        )
        sbom_names = {sbom.id: sbom.sbom_name for sbom in sboms}
        for component in components:
            key = _component_identity(component)
            unique_type, unique_value = _unique_id(component)
            aggregate = aggregates.get(key)
            if aggregate is None:
                aggregate = _ComponentAggregate(
                    key=key,
                    name=_first_text(component.name, component.normalized_name),
                    version=_first_text(component.version, component.normalized_version),
                    supplier=_first_text(component.supplier, component.normalized_supplier),
                    unique_id_type=unique_type,
                    unique_id_value=unique_value,
                    component_type=_first_text(component.component_type, "Library"),
                    origin_category=_origin_category(component),
                    license=_clean(component.license),
                    support_level=_support_level(component),
                    eos_date=_parse_date(component.eos_date),
                    eol_date=_parse_date(component.eol_date or component.eof_date),
                    patch_mechanism=_first_text(component.recommended_version, component.latest_supported_version),
                    crypto_info="",
                )
                aggregates[key] = aggregate
            aggregate.source_sboms.add(sbom_names.get(component.sbom_id, f"SBOM #{component.sbom_id}"))
            aggregate.dependencies.update(dependencies.get((component.sbom_id, component.bom_ref or ""), set()))
            aggregate.vulnerabilities.update(findings_by_component.get(component.id, set()))
            if _clean(component.lifecycle_recommendation):
                aggregate.lifecycle_recommendations.add(_clean(component.lifecycle_recommendation))
            aggregate.components.append(component)

        return sorted(aggregates.values(), key=lambda item: (item.name.casefold(), item.version.casefold(), item.key))

    def _vulnerability_rows(self, runs: dict[int, AnalysisRun]) -> list[dict[str, Any]]:
        if not runs:
            return []
        findings = list(
            self.db.execute(
                select(AnalysisFinding)
                .where(AnalysisFinding.analysis_run_id.in_([run.id for run in runs.values()]))
                .order_by(AnalysisFinding.vuln_id.asc(), AnalysisFinding.component_name.asc(), AnalysisFinding.id.asc())
            ).scalars()
        )
        component_ids = [finding.component_id for finding in findings if finding.component_id is not None]
        components = {
            component.id: component
            for component in self.db.execute(select(SBOMComponent).where(SBOMComponent.id.in_(component_ids))).scalars()
        }
        vex_rows = list(
            self.db.execute(
                select(VexStatement).where(VexStatement.sbom_id.in_([run.sbom_id for run in runs.values()]))
            ).scalars()
        )
        vex_by_key: dict[tuple[int | None, str], VexStatement] = {}
        for row in vex_rows:
            for vuln in (row.vulnerability_id, row.cve_id):
                if _clean(vuln):
                    vex_by_key[(row.component_id, _clean(vuln).casefold())] = row
                    vex_by_key[(None, _clean(vuln).casefold())] = row

        rows: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        for finding in findings:
            component = components.get(finding.component_id) if finding.component_id else None
            identity = _component_identity(component) if component else f"finding-component:{finding.component_name}:{finding.component_version}"
            key = (identity, finding.vuln_id.casefold())
            if key in seen:
                continue
            seen.add(key)
            vex = vex_by_key.get((finding.component_id, finding.vuln_id.casefold())) or vex_by_key.get(
                (None, finding.vuln_id.casefold())
            )
            status = _first_text(getattr(vex, "status", None), "Under Investigation")
            rows.append(
                {
                    "component_name": _first_text(component.name if component else None, finding.component_name),
                    "component_version": _first_text(component.version if component else None, finding.component_version),
                    "vulnerability_id": finding.vuln_id,
                    "cvss_version": _first_text(finding.cvss_version, "3.1" if _clean(finding.vector).startswith("CVSS:3.") else ""),
                    "cvss_score": finding.score,
                    "severity": _severity(finding.severity),
                    "vex_status": status,
                    "vex_justification": _first_text(getattr(vex, "justification", None)),
                    "patient_impact": "",
                    "remediation": _first_text(
                        getattr(vex, "action_statement", None),
                        getattr(vex, "mitigation", None),
                        ", ".join(_as_list(finding.fixed_versions)),
                    ),
                    "target_date": None,
                    "status": "Resolved" if status.casefold() == "fixed" else "Open",
                    "notes": _first_text(finding.source, finding.reference_url, finding.match_reason),
                }
            )
        return rows

    @staticmethod
    def _lifecycle_rows(component_rows: list[_ComponentAggregate]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for aggregate in component_rows:
            representative = aggregate.components[0] if aggregate.components else None
            if representative is None:
                continue
            status = _clean(representative.lifecycle_status)
            has_lifecycle_signal = any(
                [
                    aggregate.eos_date,
                    aggregate.eol_date,
                    status and status.casefold() not in {"unknown", "active", "supported"},
                    representative.unsupported,
                    representative.deprecated,
                    representative.is_deprecated,
                ]
            )
            if not has_lifecycle_signal:
                continue
            rows.append(
                {
                    "name": aggregate.name,
                    "version": aggregate.version,
                    "supplier": aggregate.supplier,
                    "support_level": aggregate.support_level,
                    "eos_date": aggregate.eos_date,
                    "eol_date": aggregate.eol_date,
                    "risk": _risk_for_lifecycle(representative),
                    "plan": _first_text(
                        "; ".join(sorted(aggregate.lifecycle_recommendations)),
                        representative.recommended_version and f"Upgrade to {representative.recommended_version}",
                    ),
                    "controls": "",
                    "owner": "",
                    "target_date": None,
                    "status": "Planned",
                }
            )
        return rows

    def _build_workbook(
        self,
        metadata: Fda510kReportMetadata,
        component_rows: list[_ComponentAggregate],
        vulnerability_rows: list[dict[str, Any]],
        lifecycle_rows: list[dict[str, Any]],
    ) -> bytes:
        workbook = load_workbook(self.template_path)
        if tuple(workbook.sheetnames) != EXPECTED_SHEETS:
            raise Fda510kReportError(f"FDA template sheet contract mismatch: {workbook.sheetnames}")

        metadata_ws = workbook["SBOM Metadata"]
        generated_at = datetime.now(UTC).replace(tzinfo=None, microsecond=0)
        metadata_values = {
            "C5": metadata.device_name,
            "C6": metadata.device_model_catalog_number,
            "C7": metadata.manufacturer_sponsor,
            "C8": metadata.submission_type or "510(k)",
            "C9": metadata.submission_number,
            "C10": metadata.product_code_regulation_number,
            "C11": metadata.device_software_version,
            "C12": metadata.top_level_primary_component,
            "C15": metadata.author_of_sbom_data,
            "C16": generated_at,
            "C17": metadata.sbom_version,
            "C18": metadata.sbom_formats_for_submission,
            "C19": metadata.sbom_generation_tool_and_version,
            "C20": metadata.primary_data_source,
            "C23": metadata.prepared_by,
            "C24": metadata.date_prepared,
            "C25": metadata.reviewed_approved_by,
            "C26": metadata.date_approved,
        }
        for cell, value in metadata_values.items():
            metadata_ws[cell] = value
        metadata_ws["C16"].number_format = "yyyy-mm-dd hh:mm"
        metadata_ws["C24"].number_format = "yyyy-mm-dd"
        metadata_ws["C26"].number_format = "yyyy-mm-dd"

        comp_ws = workbook["SBOM Components"]
        _clear_and_prepare_rows(comp_ws, start_row=3, style_row=5, last_row=max(5, 2 + len(component_rows)))
        for index, row in enumerate(component_rows, start=1):
            excel_row = index + 2
            dependency_parts = []
            if row.source_sboms:
                dependency_parts.append("Source SBOM(s): " + ", ".join(sorted(row.source_sboms)))
            dependency_parts.extend(sorted(row.dependencies))
            _set_row_values(
                comp_ws,
                excel_row,
                {
                    1: index,
                    2: row.name,
                    3: row.version,
                    4: row.supplier,
                    5: row.unique_id_type,
                    6: row.unique_id_value,
                    7: "; ".join(dependency_parts),
                    8: row.component_type,
                    9: row.origin_category,
                    10: row.license,
                    11: row.support_level,
                    12: row.eos_date,
                    13: row.eol_date,
                    16: ", ".join(sorted(row.vulnerabilities)),
                    17: row.patch_mechanism,
                    18: row.crypto_info,
                },
            )

        vuln_ws = workbook["Vulnerabilities & VEX"]
        _clear_and_prepare_rows(vuln_ws, start_row=3, style_row=4, last_row=max(4, 2 + len(vulnerability_rows)))
        for index, row in enumerate(vulnerability_rows, start=1):
            _set_row_values(
                vuln_ws,
                index + 2,
                {
                    1: index,
                    2: row["component_name"],
                    3: row["component_version"],
                    4: row["vulnerability_id"],
                    5: row["cvss_version"],
                    6: row["cvss_score"],
                    7: row["severity"],
                    8: row["vex_status"],
                    9: row["vex_justification"],
                    10: row["patient_impact"],
                    11: row["remediation"],
                    12: row["target_date"],
                    13: row["status"],
                    14: row["notes"],
                },
            )

        lifecycle_ws = workbook["Lifecycle & Support Plan"]
        _clear_and_prepare_rows(lifecycle_ws, start_row=3, style_row=4, last_row=max(4, 2 + len(lifecycle_rows)))
        for index, row in enumerate(lifecycle_rows, start=1):
            _set_row_values(
                lifecycle_ws,
                index + 2,
                {
                    1: index,
                    2: row["name"],
                    3: row["version"],
                    4: row["supplier"],
                    5: row["support_level"],
                    6: row["eos_date"],
                    7: row["eol_date"],
                    9: row["risk"],
                    10: row["plan"],
                    11: row["controls"],
                    12: row["owner"],
                    13: row["target_date"],
                    14: row["status"],
                },
            )

        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


__all__ = [
    "EXCEL_MEDIA_TYPE",
    "FDA_510K_TEMPLATE_PATH",
    "Fda510kExcelReportService",
    "Fda510kIncompleteAnalysisError",
    "Fda510kReportError",
    "Fda510kReportMetadata",
    "Fda510kSelection",
    "Fda510kTemplateMissingError",
    "INCOMPLETE_ANALYSIS_CODE",
]
