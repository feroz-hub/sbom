from __future__ import annotations

import json
from io import BytesIO
from uuid import uuid4

import pytest
from app.db import SessionLocal
from app.models import AnalysisFinding, AnalysisRun, Projects, SBOMComponent, SBOMSource
from openpyxl import load_workbook

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture()
def db(client):
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def _metadata() -> dict:
    return {
        "device_name": "Infusion Controller",
        "device_model_catalog_number": "IC-510K",
        "manufacturer_sponsor": "Acme Medical",
        "submission_type": "510(k)",
        "submission_number": "K260001",
        "product_code_regulation_number": "XYZ / 21 CFR 880.5725",
        "device_software_version": "9.4.1",
        "top_level_primary_component": "Infusion Controller Firmware",
        "author_of_sbom_data": "Build Pipeline",
        "sbom_version": "2026.07",
        "sbom_formats_for_submission": "CycloneDX / SPDX (machine-readable) + this workbook",
        "sbom_generation_tool_and_version": "Syft 1.2.3",
        "primary_data_source": "Persisted SBOM analysis results",
        "prepared_by": "Regulatory Ops",
        "date_prepared": "2026-07-03",
        "reviewed_approved_by": "Quality Lead",
        "date_approved": "2026-07-03",
    }


def _seed_ready_sbom(db):
    token = uuid4().hex[:8]
    project = Projects(project_name=f"FDA Project {token}", project_status=1)
    db.add(project)
    db.flush()
    raw = {
        "bomFormat": "CycloneDX",
        "components": [
            {
                "bom-ref": "pkg:maven/org.example/crypto-core@1.0.0",
                "name": "crypto-core",
                "version": "1.0.0",
            }
        ],
        "dependencies": [{"ref": "device-app", "dependsOn": ["pkg:maven/org.example/crypto-core@1.0.0"]}],
    }
    sbom = SBOMSource(
        sbom_name=f"FDA SBOM {token}",
        sbom_data=json.dumps(raw),
        projectid=project.id,
        sbom_version="2026.07",
        productver="9.4.1",
        product_name="Infusion Controller Firmware",
        created_by="Build Pipeline",
    )
    db.add(sbom)
    db.flush()
    component = SBOMComponent(
        sbom_id=sbom.id,
        bom_ref="pkg:maven/org.example/crypto-core@1.0.0",
        name="crypto-core",
        version="1.0.0",
        supplier="Example Software",
        purl="pkg:maven/org.example/crypto-core@1.0.0",
        normalized_purl="pkg:maven/org.example/crypto-core@1.0.0",
        component_type="Library",
        license="Apache-2.0",
        lifecycle_status="EOL",
        maintenance_status="No longer maintained",
        eos_date="2026-12-31",
        eol_date="2027-01-31",
        lifecycle_recommendation="Upgrade to crypto-core 1.2.0 before EOS.",
        lifecycle_checked_at="2026-07-03T00:00:00Z",
        lifecycle_source="Unit Test",
    )
    db.add(component)
    db.flush()
    run = AnalysisRun(
        sbom_id=sbom.id,
        project_id=project.id,
        run_status="FINDINGS",
        sbom_name=sbom.sbom_name,
        source="stored-test-data",
        started_on="2026-07-03T08:00:00Z",
        completed_on="2026-07-03T08:01:00Z",
        total_components=1,
        total_findings=1,
        high_count=1,
    )
    db.add(run)
    db.flush()
    db.add(
        AnalysisFinding(
            analysis_run_id=run.id,
            component_id=component.id,
            vuln_id="CVE-2026-12345",
            source="NVD",
            title="Stored vulnerability",
            severity="HIGH",
            score=8.1,
            vector="CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:H",
            published_on="2026-07-01T00:00:00Z",
            component_name=component.name,
            component_version=component.version,
            fixed_versions=json.dumps(["1.2.0"]),
            cpe="cpe:2.3:a:example:crypto-core:1.0.0:*:*:*:*:*:*:*",
        )
    )
    db.commit()
    return project, sbom


def _post_export(client, project_id: int, sbom_id: int):
    return client.post(
        f"/api/projects/{project_id}/reports/fda-510k-sbom/export",
        json={"selections": [{"sbom_id": sbom_id}], "metadata": _metadata()},
    )


def test_fda_510k_export_uses_template_and_removes_examples(client, db):
    project, sbom = _seed_ready_sbom(db)

    response = _post_export(client, project.id, sbom.id)

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == EXCEL_MEDIA_TYPE
    assert "FDA_510k_SBOM_Report" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == [
        "Instructions",
        "SBOM Metadata",
        "SBOM Components",
        "Vulnerabilities & VEX",
        "Lifecycle & Support Plan",
    ]
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True

    metadata = workbook["SBOM Metadata"]
    assert metadata["C5"].value == "Infusion Controller"
    assert metadata["C16"].is_date
    assert metadata["C24"].is_date

    components = workbook["SBOM Components"]
    assert components["A3"].value == 1
    assert components["B3"].value == "crypto-core"
    assert components["E3"].value == "PURL"
    assert components["F3"].value == "pkg:maven/org.example/crypto-core@1.0.0"
    assert components["L3"].is_date
    assert components["N3"].value == '=IF(L3="","",L3-TODAY())'
    assert components["O3"].value.startswith('=IF(L3="","",IF(L3<TODAY()')

    vulnerabilities = workbook["Vulnerabilities & VEX"]
    assert vulnerabilities["A3"].value == 1
    assert vulnerabilities["B3"].value == "crypto-core"
    assert vulnerabilities["D3"].value == "CVE-2026-12345"
    assert vulnerabilities["G3"].value == "High"

    lifecycle = workbook["Lifecycle & Support Plan"]
    assert lifecycle["A3"].value == 1
    assert lifecycle["B3"].value == "crypto-core"
    assert lifecycle["F3"].is_date

    text_values = [
        str(cell.value)
        for worksheet in [
            workbook["SBOM Components"],
            workbook["Vulnerabilities & VEX"],
            workbook["Lifecycle & Support Plan"],
        ]
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    joined = "\n".join(text_values)
    assert "EXAMPLE — replace with your data" not in joined
    assert "\nEX\n" not in f"\n{joined}\n"
    assert "OpenSSL" not in joined
    assert "Linux Kernel" not in joined
    assert "CVE-2024-XXXXX" not in joined


def test_fda_510k_export_returns_structured_409_for_incomplete_lifecycle(client, db):
    project, sbom = _seed_ready_sbom(db)
    component = db.query(SBOMComponent).filter(SBOMComponent.sbom_id == sbom.id).one()
    component.lifecycle_checked_at = None
    db.commit()

    response = _post_export(client, project.id, sbom.id)

    assert response.status_code == 409
    detail = response.json()["detail"]
    assert detail["code"] == "fda_510k_report_incomplete_analysis"
    assert detail["blockers"] == [
        {
            "sbom_id": sbom.id,
            "sbom_name": sbom.sbom_name,
            "analysis_type": "lifecycle",
            "status": "missing",
        }
    ]
